import pandas as pd
from keepa import Keepa
from datetime import datetime
from tqdm import tqdm
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import requests
import glob
import os

# ==== CONFIG ====
API_KEY = 'aja9g1k52mr0co8l8h3csj64io1slb3j7t6se1ic5taar1skedktm3pa95qb5862'

# Automatically find latest KeepaExport file with ProductFinder
export_files = glob.glob("KeepaExport-*-ProductFinder.xlsx")
if not export_files:
    raise FileNotFoundError("No matching KeepaExport-(...)-ProductFinder.xlsx file found.")
EXPORT_FILE = max(export_files, key=os.path.getctime)  # Most recently modified file

TEMPLATE_FILE = 'Final_Enriched_Affiliate_Template.xlsx'
TIMESTAMP = datetime.now().strftime('%Y%m%d_%H%M')
OUTPUT_FILE = f'Final_Enriched_Affiliate_Sheet_ImagesCharts_{TIMESTAMP}.xlsx'
DOMAIN = 'JP'
KEEPA_DOMAIN = 5  # Japan

# ==== STEP 1: LOAD FILES ====
export_df = pd.read_excel(EXPORT_FILE)
template_df = pd.read_excel(TEMPLATE_FILE)
template_df['ASIN'] = export_df['ASIN']

# Field Mappings
MAPPINGS = {
    "Title": "Title",
    "URL: Amazon": "URL: Amazon",
    "Sales Rank: Current": "Sales Rank",
    "Reviews: Rating": "Rating",
    "Reviews: Review Count": "Review Count",
    "Buy Box: Current": "Price (\u00a5)",
    "Buy Box: Stock": "Buy Box: Stock",
    "Referral Fee %": "Referral Fee %",
    "Referral Fee based on current Buy Box price": "Referral Fee based on current Buy Box price (\u00a5)",
    "Buy Box: 1 day drop %": "Discount (%)",
    "Image": "Image",
    "One Time Coupon: Absolute": "One Time Coupon: Absolute",
    "One Time Coupon: Percentage": "One Time Coupon: Percentage",
    "One Time Coupon: Subscribe & Save %": "One Time Coupon: Subscribe & Save %",
    "Categories: Root": "Categories: Root"
}

# ==== STEP 2: MAP STATIC FIELDS ====
for src_col, tgt_col in tqdm(MAPPINGS.items(), desc="Mapping columns"):
    if src_col in export_df.columns and tgt_col in template_df.columns:
        template_df[tgt_col] = export_df[src_col]

def has_extra_discount(row):
    return "Yes" if any(pd.notna(row[col]) and str(row[col]).strip() != '' for col in [
        "One Time Coupon: Absolute",
        "One Time Coupon: Percentage",
        "One Time Coupon: Subscribe & Save %"
    ]) else ""

template_df["Extra Discount"] = template_df.apply(has_extra_discount, axis=1)
template_df["Affiliate Link"] = template_df["ASIN"].apply(
    lambda asin: f"https://www.amazon.co.jp/dp/{asin}/?tag=pkmsalechanne-22"
)

# ==== STEP 3: ENRICH FROM KEEPA ====
keepa = Keepa(API_KEY)
asins = template_df["ASIN"].dropna().astype(str).tolist()
print("Querying Keepa API...")
products = keepa.query(asins, domain=DOMAIN, stats=180)

def safe_bool(val): return bool(val) if isinstance(val, bool) else False
def safe_float(val): return float(val) if isinstance(val, (int, float)) else ""

for idx, asin in tqdm(enumerate(asins), total=len(asins), desc="Enriching via Keepa API"):
    product = next((p for p in products if p['asin'] == asin), None)
    if not product:
        continue
    stats = product.get("stats", {})
    variation_count = product.get("variationCount")
    template_df.at[idx, "isLowest"] = safe_bool(stats.get("isLowest"))
    template_df.at[idx, "isLowest90"] = safe_bool(stats.get("isLowest90"))
    template_df.at[idx, "buyBox30dDropPercent"] = safe_float(stats.get("buyBox30dDropPercentage"))
    sr30 = stats.get("salesRank30")
    sr90 = stats.get("salesRank90")
    if isinstance(sr30, (int, float)) and isinstance(sr90, (int, float)):
        trend = "\u2191 Improving" if sr30 < sr90 else "\u2193 Worsening" if sr30 > sr90 else "\u2192 Stable"
    else:
        trend = "N/A"
    template_df.at[idx, "Trending BSR"] = trend
    if variation_count is not None:
        template_df.at[idx, "Variation Simplicity"] = (
            "\u2705 Simple" if variation_count <= 3 else
            "\u26a0\ufe0f Moderate" if variation_count <= 10 else
            "\u274c Complex"
        )
    else:
        template_df.at[idx, "Variation Simplicity"] = "N/A"
    drop_percent = safe_float(stats.get("buyBox30dDropPercentage"))
    fake_flag = not stats.get("isLowest90", False) and drop_percent >= 30
    template_df.at[idx, "Fake Drop Flag"] = fake_flag
    try:
        discount = safe_float(template_df.at[idx, "Discount (%)"])
        rank = int(template_df.at[idx, "Sales Rank"])
        grade = "F"
        if pd.isna(discount) or pd.isna(rank):
            grade = "F"
        elif discount >= 50 and rank <= 5000: grade = "A"
        elif discount >= 40 and rank <= 10000: grade = "B"
        elif discount >= 30 and rank <= 20000: grade = "C"
        elif discount >= 20 or rank <= 50000: grade = "D"
        if fake_flag and grade in ["A", "B", "C", "D"]:
            grade = chr(ord(grade) + 1)
        template_df.at[idx, "Grade"] = grade
    except:
        template_df.at[idx, "Grade"] = "F"

# ==== STEP 4: WRITE TO EXCEL & FORMAT ====
template_df.to_excel(OUTPUT_FILE, index=False)
wb = openpyxl.load_workbook(OUTPUT_FILE)
ws = wb.active

# Copy column widths and row heights from template
template_wb = openpyxl.load_workbook(TEMPLATE_FILE)
template_ws = template_wb.active
template_headers = {cell.value: idx for idx, cell in enumerate(template_ws[1])}

for col_idx, cell in enumerate(ws[1]):
    header = cell.value
    if header in template_headers:
        template_letter = get_column_letter(template_headers[header] + 1)
        current_letter = get_column_letter(col_idx + 1)
        template_dim = template_ws.column_dimensions.get(template_letter)
        if template_dim and template_dim.width:
            ws.column_dimensions[current_letter].width = template_dim.width

# Row heights
for row_idx in range(1, ws.max_row + 1):
    ws.row_dimensions[row_idx].height = 30 if row_idx == 1 else 230

# Wrap headers and Title
for cell in ws[1]:
    cell.alignment = Alignment(wrap_text=True)
if "Title" in template_df.columns:
    title_col = template_df.columns.get_loc("Title") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[title_col - 1].alignment = Alignment(wrap_text=True)

# Clickable URLs
if "URL: Amazon" in template_df.columns:
    url_idx = template_df.columns.get_loc("URL: Amazon") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[url_idx - 1]
        if isinstance(cell.value, str) and cell.value.startswith("http"):
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

# Format % columns
for col_name in ["Discount (%)", "Referral Fee %", "One Time Coupon: Percentage"]:
    if col_name in template_df.columns:
        col_idx = template_df.columns.get_loc(col_name) + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell = row[col_idx - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00%'

# ==== STEP 5: EMBED IMAGES & CHARTS ====
def get_col_letter(header):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            return ws.cell(row=1, column=col).column_letter
    return None

image_col_letter = get_col_letter("Image")
chart_col_letter = get_col_letter("Chart")

for idx, row in tqdm(template_df.iterrows(), total=len(template_df), desc="Embedding images and charts"):
    asin = row.get("ASIN", "")
    if not asin or pd.isna(asin): continue
    row_excel = idx + 2
    img_url = row.get("Image")
    if isinstance(img_url, str) and img_url.startswith("http"):
        try:
            img_resp = requests.get(img_url, timeout=10)
            if img_resp.status_code == 200:
                img = XLImage(BytesIO(img_resp.content))
                img.width, img.height = 200, 200
                ws.add_image(img, f"{image_col_letter}{row_excel}")
        except Exception as e:
            print(f"\u274c Image fail for {asin}: {e}")
    try:
        chart_url = f"https://api.keepa.com/graphimage?key={API_KEY}&domain={KEEPA_DOMAIN}&asin={asin}&salesrank=1&price=1&buybox=1"
        chart_resp = requests.get(chart_url, timeout=10)
        if chart_resp.status_code == 200:
            chart = XLImage(BytesIO(chart_resp.content))
            chart.width, chart.height = 565, 300
            ws.add_image(chart, f"{chart_col_letter}{row_excel}")
    except Exception as e:
        print(f"\u274c Chart fail for {asin}: {e}")

# ==== SAVE ====
wb.save(OUTPUT_FILE)
print(f"\u2705 Done. File saved to: {OUTPUT_FILE}")
