import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# === CONFIG ===
INPUT_FILE = "Final_Enriched_Affiliate_Sheet_ImagesCharts_20250613_2115.xlsx"
OUTPUT_FILE = f"Filtered_NoImages_ApprovedOnly_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

# === LOAD ===
wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb.active

# === Identify Column Letters ===
def get_col_letter_by_header(header):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            return get_column_letter(col)
    return None

def get_col_index_by_header(header):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            return col
    return None

image_col = get_col_letter_by_header("Image")
chart_col = get_col_letter_by_header("Chart")
approved_col_idx = get_col_index_by_header("Marketing Approved")

if not approved_col_idx:
    raise Exception("Marketing Approved column not found.")

# === REMOVE IMAGES ONLY IN TARGET COLUMNS ===
target_cols = {image_col, chart_col}
remaining_images = []
removed_img_count = 0

for img in ws._images:
    if hasattr(img.anchor, '_from'):
        col_letter = get_column_letter(img.anchor._from.col + 1)
        if col_letter in target_cols:
            removed_img_count += 1
            continue  # Skip = remove
    remaining_images.append(img)

ws._images = remaining_images

# === DELETE ROWS WITHOUT APPROVAL (bottom-up to avoid index shift) ===
deleted_row_count = 0
for row in range(ws.max_row, 1, -1):  # Skip header row (1)
    cell_val = ws.cell(row=row, column=approved_col_idx).value
    if str(cell_val).strip().lower() != "yes":
        ws.delete_rows(row)
        deleted_row_count += 1

# === SAVE ===
wb.save(OUTPUT_FILE)
print(f"✅ Removed {removed_img_count} images from 'Image' and 'Chart' columns.")
print(f"✅ Deleted {deleted_row_count} rows without 'Yes' in 'Marketing Approved'.")
print(f"📁 Saved to: {OUTPUT_FILE}")
