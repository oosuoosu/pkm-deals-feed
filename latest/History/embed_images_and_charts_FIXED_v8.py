import os
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from datetime import datetime
from tqdm import tqdm
import pandas as pd

# === CONFIG ===
INPUT_FILE = "Filtered_NoImages_ApprovedOnly_20250612_0043.xlsx"
OUTPUT_FILE = f"Final_Enriched_Affiliate_Sheet_ImagesCharts_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
API_KEY = "aja9g1k52mr0co8l8h3csj64io1slb3j7t6se1ic5taar1skedktm3pa95qb5862"
KEEPA_DOMAIN = 5  # Japan

# === LOAD ===
df = pd.read_excel(INPUT_FILE)
wb = load_workbook(INPUT_FILE)
ws = wb.active

# === Get Column Index by Header Name ===
def get_col_letter(header_name):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header_name:
            return ws.cell(row=1, column=col).column_letter
    return None

image_col_letter = get_col_letter("Image")
chart_col_letter = get_col_letter("Chart")
asin_col_index = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "ASIN":
        asin_col_index = col
        break

if not image_col_letter or not chart_col_letter or not asin_col_index:
    raise Exception("Missing one of the required columns: ASIN, Image, Chart.")

# === Embed ===
print("🖼️ Embedding product images and Keepa charts...")

for idx, row in tqdm(df.iterrows(), total=len(df)):
    asin = row.get("ASIN", "")
    if not asin or pd.isna(asin):
        continue

    row_excel = idx + 2
    ws.row_dimensions[row_excel].height = 215

    # === Product Image ===
    img_url = row.get("Image")
    if isinstance(img_url, str) and img_url.startswith("http"):
        try:
            resp = requests.get(img_url, timeout=10)
            if resp.status_code == 200:
                img = XLImage(BytesIO(resp.content))
                img.width = 200
                img.height = 200
                ws.add_image(img, f"{image_col_letter}{row_excel}")
        except Exception as e:
            print(f"❌ Image fail for ASIN {asin}: {e}")

    # === Chart Image (Keepa) ===
    try:
        chart_url = f"https://api.keepa.com/graphimage?key={API_KEY}&domain={KEEPA_DOMAIN}&asin={asin}&salesrank=1&price=1&buybox=1"
        chart_resp = requests.get(chart_url, timeout=10)
        if chart_resp.status_code == 200:
            chart_img = XLImage(BytesIO(chart_resp.content))
            chart_img.width = 565
            chart_img.height = 300
            ws.add_image(chart_img, f"{chart_col_letter}{row_excel}")
    except Exception as e:
        print(f"❌ Chart fail for ASIN {asin}: {e}")

# === Set Header Row Height ===
ws.row_dimensions[1].height = 30

# === Save ===
wb.save(OUTPUT_FILE)
print(f"✅ Done. File saved to: {OUTPUT_FILE}")
