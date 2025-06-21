import os
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from datetime import datetime
from tqdm import tqdm
import pandas as pd

# === CONFIG ===
INPUT_FILE = "Final_Enriched_Affiliate_Sheet_Styled_20250611_2004.xlsx"
OUTPUT_FILE = f"Final_Enriched_Affiliate_Sheet_ImagesCharts_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
API_KEY = "aja9g1k52mr0co8l8h3csj64io1slb3j7t6se1ic5taar1skedktm3pa95qb5862"
KEEPA_DOMAIN = 5  # Japan

# === LOAD ===
df = pd.read_excel(INPUT_FILE)
wb = load_workbook(INPUT_FILE)
ws = wb.active

# === Get Column Index ===
def get_col_index(header):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            return col
    return None

image_col = get_col_index("Image")
asin_col = get_col_index("ASIN")
chart_col = get_col_index("Chart") or 2  # fallback to column B
if not asin_col or not image_col:
    raise Exception("Missing ASIN or Image column.")

# === PROCESS ===
print("🖼️ Embedding product images and Keepa charts...")
for idx, row in tqdm(df.iterrows(), total=len(df)):
    asin = row.get("ASIN", "")
    if not asin or pd.isna(asin):
        continue

    row_excel = idx + 2

    # === Product Image ===
    img_url = row.get("Image")
    if isinstance(img_url, str) and img_url.startswith("http"):
        try:
            resp = requests.get(img_url, timeout=10)
            if resp.status_code == 200:
                img = XLImage(BytesIO(resp.content))
                img.width = 80
                img.height = 80
                ws.add_image(img, f"A{row_excel}")
                ws.row_dimensions[row_excel].height = 100
        except Exception as e:
            print(f"❌ Image fail for ASIN {asin}: {e}")

    # === Keepa Chart Image (Pre-rendered) ===
    try:
        chart_url = f"https://api.keepa.com/graphimage?key={API_KEY}&domain={KEEPA_DOMAIN}&asin={asin}&salesrank=1&price=1&buybox=1"
        chart_resp = requests.get(chart_url, timeout=10)
        if chart_resp.status_code == 200:
            chart_img = XLImage(BytesIO(chart_resp.content))
            chart_img.width = 120
            chart_img.height = 80
            chart_cell = f"B{row_excel}"
            ws.add_image(chart_img, chart_cell)
    except Exception as e:
        print(f"❌ Chart fail for ASIN {asin}: {e}")

# === SAVE ===
wb.save(OUTPUT_FILE)
print(f"✅ Done. File saved to: {OUTPUT_FILE}")
