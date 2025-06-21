
import os
import requests
import pandas as pd
from io import BytesIO
from datetime import datetime
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# === CONFIG ===
API_KEY = "aja9g1k52mr0co8l8h3csj64io1slb3j7t6se1ic5taar1skedktm3pa95qb5862"
KEEPA_DOMAIN = 5  # Japan
IMAGE_FOLDER = "affiliate_images"

# === LOAD ===
df = pd.read_excel(INPUT_FILE)
wb = load_workbook(INPUT_FILE)
ws = wb.active

# === Get Column Helpers ===
def get_col_letter_by_header(header):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            return get_column_letter(col)
    return None

def get_col_index_by_header(header):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            return col
    return None

image_col = get_col_letter_by_header("Image")
approved_col_idx = get_col_index_by_header("Marketing Approved")
asin_col_index = get_col_index_by_header("ASIN")


# === REMOVE OLD IMAGES IN TARGET COLUMNS ===
remaining_images = []
removed_img_count = 0
for img in ws._images:
    if hasattr(img.anchor, '_from'):
        col_letter = get_column_letter(img.anchor._from.col + 1)
        if col_letter in target_cols:
            removed_img_count += 1
            continue
    remaining_images.append(img)
ws._images = remaining_images

# === DELETE UNAPPROVED ROWS ===
deleted_row_count = 0
for row in range(ws.max_row, 1, -1):
    cell_val = ws.cell(row=row, column=approved_col_idx).value
    if str(cell_val).strip().lower() != "yes":
        ws.delete_rows(row)
        deleted_row_count += 1

# === Refresh DataFrame ===
df = pd.read_excel(INPUT_FILE)
df = df[df["Marketing Approved"].astype(str).str.lower().str.strip() == "yes"].reset_index(drop=True)

# === PREP IMAGE FOLDER STRUCTURE ===
os.makedirs(IMAGE_FOLDER, exist_ok=True)

# === EMBED IMAGES & SAVE TO FOLDERS ===
for idx, row in tqdm(df.iterrows(), total=len(df), desc="Processing rows"):
    asin = row.get("ASIN", "")
    if not asin or pd.isna(asin):
        continue
    row_excel = idx + 2
    ws.row_dimensions[row_excel].height = 215


    # Product Image
    img_url = row.get("Image")
    if isinstance(img_url, str) and img_url.startswith("http"):
        try:
            resp = requests.get(img_url, timeout=10)
            if resp.status_code == 200:
                img = XLImage(BytesIO(resp.content))
                img.width = 200
                img.height = 200
                ws.add_image(img, f"{image_col}{row_excel}")
                    f.write(resp.content)
        except Exception as e:
            print(f"❌ Image fail for ASIN {asin}: {e}")

    try:
    except Exception as e:

# Header height
ws.row_dimensions[1].height = 30

# === SAVE ===
wb.save(OUTPUT_FILE)
print(f"✅ Removed {removed_img_count} old images.")
print(f"✅ Deleted {deleted_row_count} unapproved rows.")
print(f"📁 Saved enriched file to: {OUTPUT_FILE}")