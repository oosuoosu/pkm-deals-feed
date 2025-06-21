
import os
import requests
import pandas as pd
from io import BytesIO
from datetime import datetime
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# === CONFIG ===
FOLDER = "C:/Users/Philip Work/Documents/Python"
IMAGE_FOLDER = os.path.join(FOLDER, "affiliate_images")
MARKETING_FOLDER = os.path.join(FOLDER, "Marketing")
PREFIX = "Final_Enriched_Affiliate_Sheet_ImagesCharts_"
API_KEY = "aja9g1k52mr0co8l8h3csj64io1slb3j7t6se1ic5taar1skedktm3pa95qb5862"
DOMAIN = 5  # Japan

# === Ensure Marketing folder exists ===
os.makedirs(MARKETING_FOLDER, exist_ok=True)

# === Find latest input file ===
files = [f for f in os.listdir(FOLDER) if f.startswith(PREFIX) and f.endswith(".xlsx")]
if not files:
    raise FileNotFoundError(f"No Excel file found in {FOLDER} with prefix '{PREFIX}'")

latest_file = max(files, key=lambda f: os.path.getmtime(os.path.join(FOLDER, f)))
INPUT_FILE = os.path.join(FOLDER, latest_file)
print(f"📂 Found latest file: {INPUT_FILE}")

OUTPUT_FILE = os.path.join(
    FOLDER,
    f"Final_Enriched_Affiliate_ImagesOnly_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
)

# === LOAD ===
df = pd.read_excel(INPUT_FILE)
wb = load_workbook(INPUT_FILE)
ws = wb.active

# === Get Column Helpers ===
def get_col_letter_by_header(header):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            return get_column_letter(col)
    return None

def get_col_index_by_header(header):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            return col
    return None

image_col = get_col_letter_by_header("Image")
approved_col_idx = get_col_index_by_header("Marketing Approved")
asin_col_index = get_col_index_by_header("ASIN")

if not approved_col_idx or not image_col or not asin_col_index:
    raise Exception("Missing required column(s): ASIN, Image, or Marketing Approved")

# === REMOVE OLD IMAGES ===
remaining_images = []
removed_img_count = 0
for img in ws._images:
    if hasattr(img.anchor, '_from'):
        col_letter = get_column_letter(img.anchor._from.col + 1)
        if col_letter == image_col:
            removed_img_count += 1
            continue
    remaining_images.append(img)
ws._images = remaining_images

# === DELETE UNAPPROVED ROWS ===
deleted_row_count = 0
for row in range(ws.max_row, 1, -1):
    val = ws.cell(row=row, column=approved_col_idx).value
    if str(val).strip().lower() != "yes":
        ws.delete_rows(row)
        deleted_row_count += 1

# === Refresh Approved Rows ===
df = pd.read_excel(INPUT_FILE)
df = df[df["Marketing Approved"].astype(str).str.lower().str.strip() == "yes"].reset_index(drop=True)

# === PREP IMAGE FOLDER ===
os.makedirs(IMAGE_FOLDER, exist_ok=True)

# === Embed Images & Download Keepa Marketing Images ===
print("🖼️ Embedding and saving product images...")
for idx, row in tqdm(df.iterrows(), total=len(df), desc="Processing rows"):
    asin = row.get("ASIN", "")
    if not asin or pd.isna(asin):
        continue
    row_excel = idx + 2
    ws.row_dimensions[row_excel].height = 215

    # Embed primary image into Excel and save it as <ASIN>.jpg
    img_url = row.get("Image")
    if isinstance(img_url, str) and img_url.startswith("http"):
        try:
            resp = requests.get(img_url, timeout=10)
            if resp.status_code == 200:
                img = XLImage(BytesIO(resp.content))
                img.width = 200
                img.height = 200
                ws.add_image(img, f"{image_col}{row_excel}")
                img_path = os.path.join(IMAGE_FOLDER, f"{asin}.jpg")
                with open(img_path, "wb") as f:
                    f.write(resp.content)
        except Exception as e:
            print(f"❌ Failed to embed/save image for {asin}: {e}")

    # Download additional marketing images from Keepa
    try:
        keepa_url = f"https://api.keepa.com/product?key={API_KEY}&domain={DOMAIN}&asin={asin}"
        r = requests.get(keepa_url, timeout=15)
        r.raise_for_status()
        data = r.json()
        product = data.get("products", [{}])[0]
        images = product.get("imagesCSV", "")
        if images:
            urls = images.split(",")
            for i, suffix in enumerate(urls):
                full_url = f"https://images-na.ssl-images-amazon.com/images/I/{suffix}"
                try:
                    img_resp = requests.get(full_url, timeout=10)
                    if img_resp.status_code == 200:
                        img_file = os.path.join(MARKETING_FOLDER, f"{asin}_{i+1}.jpg")
                        with open(img_file, "wb") as out:
                            out.write(img_resp.content)
                except Exception as e:
                    print(f"⚠️ Could not save image {i+1} for {asin}: {e}")
    except Exception as e:
        print(f"⚠️ Failed Keepa API call for {asin}: {e}")

# Header row height
ws.row_dimensions[1].height = 30

# === SAVE WORKBOOK ===
wb.save(OUTPUT_FILE)
print(f"✅ Removed {removed_img_count} old images.")
print(f"✅ Deleted {deleted_row_count} unapproved rows.")
print(f"📁 Saved to: {OUTPUT_FILE}")
