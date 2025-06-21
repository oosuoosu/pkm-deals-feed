import os
import io
import requests
import pandas as pd
from datetime import datetime, timedelta
from keepa import Keepa
from tqdm import tqdm
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

# --- CONFIG ---
API_KEY = "aja9g1k52mr0co8l8h3csj64io1slb3j7t6se1ic5taar1skedktm3pa95qb5862"
INPUT_FILE = "Final_Enriched_Affiliate_Sheet_Styled_20250611_1127.xlsx"
OUTPUT_FILE = f"Final_Enriched_Affiliate_Sheet_Embedded_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

keepa = Keepa(API_KEY)

# --- Chart Generation ---
def fetch_and_plot_chart(asin):
    try:
        product = keepa.query(asin, domain="JP", history=True)[0]
        price_data = product.get("csv", [])[0]

        base_date = datetime(2011, 1, 1)
        price_points = []

        for entry in price_data:
            if isinstance(entry, list) and len(entry) >= 2:
                minutes_since_base = entry[0]
                price_val = entry[1]
            elif isinstance(entry, list) and len(entry) == 2:
                minutes_since_base, price_val = entry
            else:
                continue
            if price_val <= 0:
                continue
            timestamp = base_date + timedelta(minutes=minutes_since_base)
            price = price_val / 100
            price_points.append((timestamp, price))

        two_months_ago = datetime.now() - timedelta(days=60)
        filtered = [(t, p) for t, p in price_points if t >= two_months_ago]

        if len(filtered) < 2:
            return None

        x, y = zip(*filtered)
        plt.figure(figsize=(4, 2))
        plt.plot(x, y, linewidth=1.5)
        plt.xticks(rotation=45)
        plt.tight_layout()

        buf = io.BytesIO()
        plt.savefig(buf, format="png", dpi=100)
        plt.close()
        buf.seek(0)
        return buf
    except Exception as e:
        print(f"❌ Chart error for {asin}: {e}")
        return None

# --- Main Embed Script ---
def main():
    df = pd.read_excel(INPUT_FILE)
    wb = load_workbook(INPUT_FILE)
    ws = wb.active

    asin_col = list(df.columns).index("ASIN") + 1
    img_col = list(df.columns).index("Image") + 1
    chart_col = list(df.columns).index("Chart") + 1

    for idx, row in tqdm(df.iterrows(), total=len(df), desc="Embedding"):
        excel_row = idx + 2
        asin = str(row.get("ASIN", "")).strip()

        # --- Image embedding ---
        img_url = row.get("Image", "")
        if img_url:
            try:
                img_data = requests.get(img_url, timeout=5).content
                img_stream = io.BytesIO(img_data)
                img = ExcelImage(img_stream)
                img.width, img.height = 100, 100
                img.anchor = ws.cell(row=excel_row, column=img_col).coordinate
                ws.add_image(img)
            except Exception as e:
                print(f"❌ Image error for {asin}: {e}")

        # --- Chart embedding ---
        chart_buf = fetch_and_plot_chart(asin)
        if chart_buf:
            try:
                chart_img = ExcelImage(chart_buf)
                chart_img.width, chart_img.height = 200, 100
                chart_img.anchor = ws.cell(row=excel_row, column=chart_col).coordinate
                ws.add_image(chart_img)
                ws.row_dimensions[excel_row].height = 80
            except Exception as e:
                print(f"❌ Chart embed error for {asin}: {e}")

    wb.save(OUTPUT_FILE)
    print(f"✅ Embeds complete → {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
