import pandas as pd
import openpyxl
from openpyxl.styles import numbers
from datetime import datetime
import pytz
import os

# === CONFIGURATION ===
TEMPLATE_FILE = "Final_Enriched_Affiliate_Template.xlsx"
INPUT_FILE = "Final_Enriched_Affiliate_Sheet_Formatted_20250611_0024.xlsx"
OUTPUT_FILE = f"Final_Enriched_Affiliate_Sheet_Styled_{datetime.now(pytz.timezone('America/Vancouver')).strftime('%Y%m%d_%H%M')}.xlsx"

# === STEP 1: COPY TEMPLATE FORMATTING ===
template_wb = openpyxl.load_workbook(TEMPLATE_FILE)
template_ws = template_wb.active

row_heights = {row: template_ws.row_dimensions[row].height for row in range(1, template_ws.max_row + 1) if template_ws.row_dimensions[row].height}
col_widths = {col: template_ws.column_dimensions[col].width for col in template_ws.column_dimensions}

# === STEP 2: LOAD ENRICHED FILE ===
df = pd.read_excel(INPUT_FILE)
df.to_excel(OUTPUT_FILE, index=False)

# === STEP 3: APPLY FORMATTING ===
wb = openpyxl.load_workbook(OUTPUT_FILE)
ws = wb.active

# Set row heights (and force Excel to respect them by inserting a space in each row)
for row, height in row_heights.items():
    ws.row_dimensions[row].height = height
    if not ws.cell(row=row, column=1).value:
        ws.cell(row=row, column=1, value=" ")  # Trigger Excel to respect height

# Set column widths
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Format percentage columns
percent_columns = ["Discount (%)", "Referral Fee %"]
header = [cell.value for cell in ws[1]]

for col_name in percent_columns:
    if col_name in header:
        col_idx = header.index(col_name) + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            try:
                val = float(str(cell.value).replace("%", "")) / 100 if isinstance(cell.value, str) else float(cell.value)
                cell.value = val
                cell.number_format = '0.0%'
            except:
                continue

# Save final styled version
wb.save(OUTPUT_FILE)
print(f"✅ Styled file saved to {OUTPUT_FILE}")
