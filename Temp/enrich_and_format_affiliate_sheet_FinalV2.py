import pandas as pd
from keepa import Keepa
from datetime import datetime
from tqdm import tqdm
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

# ==== CONFIG ====
API_KEY = 'aja9g1k52mr0co8l8h3csj64io1slb3j7t6se1ic5taar1skedktm3pa95qb5862'
EXPORT_FILE = 'KeepaExport-2025-ProductFinder.xlsx'
TEMPLATE_FILE = 'Final_Enriched_Affiliate_Template.xlsx'
TIMESTAMP = datetime.now().strftime('%Y%m%d_%H%M')
OUTPUT_FILE = f'Final_Enriched_Affiliate_Sheet_Styled_{TIMESTAMP}.xlsx'
DOMAIN = 'JP'

# ==== STEP 1: LOAD FILES ====
export_df = pd.read_excel(EXPORT_FILE)
template_df = pd.read_excel(TEMPLATE_FILE)

# Fill ASINs from export into template
template_df['ASIN'] = export_df['ASIN']

# Define field mappings (source -> target)
MAPPINGS = {
    "Title": "Title",
    "URL: Amazon": "URL: Amazon",
    "Sales Rank: Current": "Sales Rank",
    "Reviews: Rating": "Rating",
    "Reviews: Review Count": "Review Count",
    "Buy Box: Current": "Price (¥)",
    "Buy Box: Stock": "Buy Box: Stock",
    "Referral Fee %": "Referral Fee %",
    "Referral Fee based on current Buy Box price": "Referral Fee based on current Buy Box price (¥)",
    "Buy Box: 1 day drop %": "Discount (%)",
    "Image": "Image",
    "One Time Coupon: Absolute": "One Time Coupon: Absolute",
    "One Time Coupon: Percentage": "One Time Coupon: Percentage",
    "One Time Coupon: Subscribe & Save %": "One Time Coupon: Subscribe & Save %"
}

# ==== STEP 2: MAP STATIC FIELDS ====
for src_col, tgt_col in tqdm(MAPPINGS.items(), desc="Mapping columns"):
    if src_col in export_df.columns and tgt_col in template_df.columns:
        template_df[tgt_col] = export_df[src_col]
    else:
        print(f"⚠️ Missing column: '{src_col}' in export or '{tgt_col}' in template")

# Extra Discount Logic
def has_extra_discount(row):
    return "Yes" if any(pd.notna(row[col]) and str(row[col]).strip() != '' for col in [
        "One Time Coupon: Absolute",
        "One Time Coupon: Percentage",
        "One Time Coupon: Subscribe & Save %"
    ]) else ""

template_df["Extra Discount"] = template_df.apply(has_extra_discount, axis=1)

# Affiliate Link
template_df["Affiliate Link"] = template_df["ASIN"].apply(
    lambda asin: f"https://www.amazon.co.jp/dp/{asin}/?tag=pkmsalechanne-22"
)

# ==== STEP 3: API ENRICHMENT ====
keepa = Keepa(API_KEY)
asins = template_df["ASIN"].dropna().astype(str).tolist()

print("Querying Keepa API...")
products = keepa.query(asins, domain=DOMAIN, stats=180)

def safe_bool(val):
    return bool(val) if isinstance(val, bool) else False

def safe_float(val):
    try:
        return float(val)
    except:
        return ""

for idx, asin in tqdm(enumerate(asins), total=len(asins), desc="Enriching via Keepa API"):
    product = next((p for p in products if p['asin'] == asin), None)
    if not product:
        continue

    stats = product.get("stats", {})
    variation_count = product.get("variationCount")

    template_df.at[idx, "isLowest"] = safe_bool(stats.get("isLowest", False))
    template_df.at[idx, "isLowest90"] = safe_bool(stats.get("isLowest90", False))
    template_df.at[idx, "buyBox30dDropPercent"] = safe_float(stats.get("buyBox30dDropPercentage"))

    # Trending BSR
    sr30 = stats.get("salesRank30")
    sr90 = stats.get("salesRank90")
    if sr30 and sr90:
        trend = "↑ Improving" if sr30 < sr90 else "↓ Worsening" if sr30 > sr90 else "→ Stable"
    else:
        trend = "N/A"
    template_df.at[idx, "Trending BSR"] = trend

    # Variation Simplicity
    if variation_count is not None:
        if variation_count <= 3:
            var_label = "✅ Simple"
        elif variation_count <= 10:
            var_label = "⚠️ Moderate"
        else:
            var_label = "❌ Complex"
    else:
        var_label = "N/A"
    template_df.at[idx, "Variation Simplicity"] = var_label

    # Fake Drop Flag
    drop_percent = safe_float(stats.get("buyBox30dDropPercentage"))
    fake_flag = not stats.get("isLowest90", False) and drop_percent >= 30
    template_df.at[idx, "Fake Drop Flag"] = fake_flag

    # Grade
    try:
        discount = safe_float(template_df.at[idx, "Discount (%)"])
        rank = int(template_df.at[idx, "Sales Rank"])
        grade = "F"
        if pd.isna(discount) or pd.isna(rank):
            grade = "F"
        elif discount >= 50 and rank <= 5000:
            grade = "A"
        elif discount >= 40 and rank <= 10000:
            grade = "B"
        elif discount >= 30 and rank <= 20000:
            grade = "C"
        elif discount >= 20 or rank <= 50000:
            grade = "D"
        else:
            grade = "F"
        if fake_flag and grade in ["A", "B", "C", "D"]:
            grade = chr(ord(grade) + 1) if grade != "F" else "F"
        template_df.at[idx, "Grade"] = grade
    except:
        template_df.at[idx, "Grade"] = "F"

# ==== STEP 4: STYLE OUTPUT ====
template_df.to_excel(OUTPUT_FILE, index=False)

wb = openpyxl.load_workbook(OUTPUT_FILE)
ws = wb.active

# Load template workbook for column width/height
template_wb = openpyxl.load_workbook(TEMPLATE_FILE)
template_ws = template_wb.active

# Copy column widths
for col_idx, col in enumerate(ws.columns, start=1):
    letter = get_column_letter(col_idx)
    try:
        ws.column_dimensions[letter].width = template_ws.column_dimensions[letter].width
    except:
        pass

# Copy row heights
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    r_idx = row[0].row
    if template_ws.row_dimensions.get(r_idx) and template_ws.row_dimensions[r_idx].height:
        ws.row_dimensions[r_idx].height = template_ws.row_dimensions[r_idx].height

# Format % columns
percent_columns = ["Discount (%)", "Referral Fee %", "One Time Coupon: Percentage"]
for col_name in percent_columns:
    if col_name in template_df.columns:
        col_idx = template_df.columns.get_loc(col_name) + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell = row[col_idx - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00%'

wb.save(OUTPUT_FILE)
print(f"✅ Final styled file saved: {OUTPUT_FILE}")
