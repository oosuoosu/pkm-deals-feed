import openpyxl
from datetime import datetime
import shutil

# === CONFIG ===
TEMPLATE_FILE = "Final_Enriched_Affiliate_Template.xlsx"
INPUT_FILE = "Final_Enriched_Affiliate_Sheet_20250611_0001.xlsx"  # <-- Update this per run
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M")
OUTPUT_FILE = f"Final_Enriched_Affiliate_Sheet_Formatted_{TIMESTAMP}.xlsx"

# === Copy to preserve original ===
shutil.copy(INPUT_FILE, OUTPUT_FILE)

# === Load workbooks ===
template_wb = openpyxl.load_workbook(TEMPLATE_FILE)
template_ws = template_wb.active

output_wb = openpyxl.load_workbook(OUTPUT_FILE)
output_ws = output_wb.active

# === Apply Column Widths ===
for col_letter, col_dim in template_ws.column_dimensions.items():
    if col_letter in output_ws.column_dimensions:
        output_ws.column_dimensions[col_letter].width = col_dim.width

# === Apply Row Heights ===
for row_num, row_dim in template_ws.row_dimensions.items():
    if row_num in output_ws.row_dimensions:
        output_ws.row_dimensions[row_num].height = row_dim.height

# === Format Percent Columns ===
PERCENT_COLUMNS = ["Discount (%)", "Referral Fee %"]
header = [cell.value for cell in output_ws[1]]
for col_name in PERCENT_COLUMNS:
    if col_name in header:
        col_idx = header.index(col_name) + 1
        for row in output_ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell.value, (float, int)):
                    cell.number_format = '0.0%'

# === Save and Done ===
output_wb.save(OUTPUT_FILE)
print(f"✅ Row heights, column widths, and % formatting applied → {OUTPUT_FILE}")
