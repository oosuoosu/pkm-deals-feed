import pandas as pd
from keepa import Keepa
from datetime import datetime
from tqdm import tqdm
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# ==== CONFIG ====
API_KEY = 'aja9g1k52mr0co8l8h3csj64io1slb3j7t6se1ic5taar1skedktm3pa95qb5862'
EXPORT_FILE = 'KeepaExport-2025-ProductFinder.xlsx'
TEMPLATE_FILE = 'Final_Enriched_Affiliate_Template.xlsx'
TIMESTAMP = datetime.now().strftime('%Y%m%d_%H%M')
OUTPUT_FILE = f'Final_Enriched_Affiliate_Sheet_Styled_{TIMESTAMP}.xlsx'
DOMAIN = 'JP'

# ==== STEP 1: LOAD FILES ====
export_df = pd.read_excel(EXPORT_FILE)
template_df = pd.read_excel(TEMPLATE_FILE)
template_df['ASIN'] = export_df['ASIN']

MAPPINGS = {
    "Title": "Title",
    "URL: Amazon": "URL: Amazon",
    "Sales Rank: Current": "Sales Rank",
    "Reviews: Rating": "Rating",
    "Reviews: Review Count": "Review Count",
    "Buy Box: Current": "Price (¥)",
    "Buy Box: Stock": "Buy Box: Stock",
    "Referral Fee %": "Referral Fee %",
    "Referral Fee based on current Buy Box price": "Referral Fee based on current Buy Box price (¥)",
    "Buy Box: 1 day drop %": "Discount (%)",
    "Image": "Image",
    "One Time Coupon: Absolute": "One Time Coupon: Absolute",
    "One Time Coupon: Percentage": "One Time Coupon: Percentage",
    "One Time Coupon: Subscribe & Save %": "One Time Coupon: Subscribe & Save %",
    "Categories: Root": "Categories: Root"
}

# ==== STEP 2: MAP STATIC FIELDS ====
for src_col, tgt_col in tqdm(MAPPINGS.items(), desc="Mapping columns"):
    if src_col in export_df.columns and tgt_col in template_df.columns:
        template_df[tgt_col] = export_df[src_col]
    else:
        print(f"⚠️ Missing column: '{src_col}' in export or '{tgt_col}' in template")

# Extra Discount Logic
def has_extra_discount(row):
    return "Yes" if any(pd.notna(row[col]) and str(row[col]).strip() != '' for col in [
        "One Time Coupon: Absolute",
        "One Time Coupon: Percentage",
        "One Time Coupon: Subscribe & Save %"
    ]) else ""

template_df["Extra Discount"] = template_df.apply(has_extra_discount, axis=1)
template_df["Affiliate Link"] = template_df["ASIN"].apply(
    lambda asin: f"https://www.amazon.co.jp/dp/{asin}/?tag=pkmsalechanne-22"
)

# ==== STEP 3: API ENRICHMENT ====
keepa = Keepa(API_KEY)
asins = template_df["ASIN"].dropna().astype(str).tolist()

print("Querying Keepa API...")
products = keepa.query(asins, domain=DOMAIN, stats=180)

def safe_bool(val):
    return bool(val) if isinstance(val, bool) else False

def safe_float(val):
    try:
        return float(val)
    except:
        return ""

for idx, asin in tqdm(enumerate(asins), total=len(asins), desc="Enriching via Keepa API"):
    product = next((p for p in products if p['asin'] == asin), None)
    if not product:
        continue

    stats = product.get("stats", {})
    variation_count = product.get("variationCount")

    template_df.at[idx, "isLowest"] = safe_bool(stats.get("isLowest", False))
    template_df.at[idx, "isLowest90"] = safe_bool(stats.get("isLowest90", False))
    template_df.at[idx, "buyBox30dDropPercent"] = safe_float(stats.get("buyBox30dDropPercentage"))

    sr30 = stats.get("salesRank30")
    sr90 = stats.get("salesRank90")
    if sr30 and sr90:
        trend = "↑ Improving" if sr30 < sr90 else "↓ Worsening" if sr30 > sr90 else "→ Stable"
    else:
        trend = "N/A"
    template_df.at[idx, "Trending BSR"] = trend

    if variation_count is not None:
        if variation_count <= 3:
            var_label = "✅ Simple"
        elif variation_count <= 10:
            var_label = "⚠️ Moderate"
        else:
            var_label = "❌ Complex"
    else:
        var_label = "N/A"
    template_df.at[idx, "Variation Simplicity"] = var_label

    drop_percent = safe_float(stats.get("buyBox30dDropPercentage"))
    fake_flag = not stats.get("isLowest90", False) and drop_percent >= 30
    template_df.at[idx, "Fake Drop Flag"] = fake_flag

    try:
        discount = safe_float(template_df.at[idx, "Discount (%)"])
        rank = int(template_df.at[idx, "Sales Rank"])
        grade = "F"
        if pd.isna(discount) or pd.isna(rank):
            grade = "F"
        elif discount >= 50 and rank <= 5000:
            grade = "A"
        elif discount >= 40 and rank <= 10000:
            grade = "B"
        elif discount >= 30 and rank <= 20000:
            grade = "C"
        elif discount >= 20 or rank <= 50000:
            grade = "D"
        else:
            grade = "F"
        if fake_flag and grade in ["A", "B", "C", "D"]:
            grade = chr(ord(grade) + 1) if grade != "F" else "F"
        template_df.at[idx, "Grade"] = grade
    except:
        template_df.at[idx, "Grade"] = "F"

# ==== STEP 4: STYLE OUTPUT ====
template_df.to_excel(OUTPUT_FILE, index=False)
wb = openpyxl.load_workbook(OUTPUT_FILE)
ws = wb.active

template_wb = openpyxl.load_workbook(TEMPLATE_FILE)
template_ws = template_wb.active

# Match column widths
header_row = next(ws.iter_rows(min_row=1, max_row=1))
template_header_row = next(template_ws.iter_rows(min_row=1, max_row=1))
template_headers = {cell.value: idx for idx, cell in enumerate(template_header_row)}

for col_idx, cell in enumerate(header_row):
    header = cell.value
    if header in template_headers:
        template_col_letter = get_column_letter(template_headers[header] + 1)
        current_col_letter = get_column_letter(col_idx + 1)
        template_dim = template_ws.column_dimensions.get(template_col_letter)
        if template_dim and template_dim.width:
            ws.column_dimensions[current_col_letter].width = template_dim.width

# Set row height explicitly
for row_idx in range(1, ws.max_row + 1):
    if row_idx == 1:
        ws.row_dimensions[row_idx].height = 30
    else:
        ws.row_dimensions[row_idx].height = 230

# Format % columns
percent_columns = ["Discount (%)", "Referral Fee %", "One Time Coupon: Percentage"]
for col_name in percent_columns:
    if col_name in template_df.columns:
        col_idx = template_df.columns.get_loc(col_name) + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell = row[col_idx - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00%'

# Wrap text in "Title" column
if "Title" in template_df.columns:
    title_idx = template_df.columns.get_loc("Title") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[title_idx - 1]
        cell.alignment = Alignment(wrap_text=True)

# Wrap text for all headers
for cell in ws[1]:
    cell.alignment = Alignment(wrap_text=True)
    
# Make "URL: Amazon" column cells clickable
if "URL: Amazon" in template_df.columns:
    url_idx = template_df.columns.get_loc("URL: Amazon") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[url_idx - 1]
        if isinstance(cell.value, str) and cell.value.startswith("http"):
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"


wb.save(OUTPUT_FILE)
print(f"✅ Final styled file saved: {OUTPUT_FILE}")
