import os
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO

# === CONFIGURATION ===
INPUT_FILE = "Final_Enriched_Affiliate_Sheet_20250610_1142.xlsx"  # Replace with your file
OUTPUT_FILE = "Final_Enriched_Affiliate_Sheet_WITH_IMAGES.xlsx"
IMAGE_WIDTH = 100
IMAGE_HEIGHT = 100

# === HEADER COLUMN NAMES ===
ASIN_COLUMN = "ASIN"
IMAGE_COLUMN = "Image"
CHART_COLUMN = "Chart"

# === FUNCTIONS ===
def get_col_index(ws, header_name):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header_name:
            return col
    raise Exception(f"Column '{header_name}' not found.")

# === MAIN LOGIC ===
wb = load_workbook(INPUT_FILE)
ws = wb.active

asin_col = get_col_index(ws, ASIN_COLUMN)
image_col = get_col_index(ws, IMAGE_COLUMN)
chart_col = get_col_index(ws, CHART_COLUMN)

for row in range(2, ws.max_row + 1):
    asin = ws.cell(row=row, column=asin_col).value
    if not asin:
        continue

    # Image URLs
    image_url = f"https://images-na.ssl-images-amazon.com/images/P/{asin}.jpg"
    chart_url = f"https://api.keepa.com/graphimage?key=aja9g1k52mr0co8l8h3csj64io1slb3j7t6se1ic5taar1skedktm3pa95qb5862&domain=5&asin={asin}&salesrank=1&price=1&buybox=1"

    for url, col in [(image_url, image_col), (chart_url, chart_col)]:
        try:
            response = requests.get(url, timeout=10)
            if response.status_code == 200:
                img = XLImage(BytesIO(response.content))
                img.width = IMAGE_WIDTH
                img.height = IMAGE_HEIGHT
                cell = f"{chr(64 + col)}{row}"
                ws.add_image(img, cell)
        except Exception as e:
            print(f"Image error on ASIN {asin}: {e}")

# Adjust column width for image/chart
for col in [image_col, chart_col]:
    ws.column_dimensions[chr(64 + col)].width = 18

# SAVE
wb.save(OUTPUT_FILE)
print(f"✅ Done: {OUTPUT_FILE}")
