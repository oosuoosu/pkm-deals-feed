import pandas as pd
from keepa import Keepa
from tqdm import tqdm
from datetime import datetime
import openpyxl
import shutil

# === CONFIG ===
API_KEY = "aja9g1k52mr0co8l8h3csj64io1slb3j7t6se1ic5taar1skedktm3pa95qb5862"
AFFILIATE_TAG = "pkmsalechanne-22"
INPUT_EXPORT = "KeepaExport-2025-ProductFinder.xlsx"
TEMPLATE_FILE = "Final_Enriched_Affiliate_Template.xlsx"
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M")
OUTPUT_FILE = f"Final_Enriched_Affiliate_Sheet_Styled_{TIMESTAMP}.xlsx"

# === COLUMN MAPPINGS ===
MAPPED_COLUMNS = {
    "Title": "Title",
    "URL: Amazon": "URL: Amazon",
    "Sales Rank: Current": "Sales Rank",
    "Reviews: Rating": "Rating",
    "Reviews: Review Count": "Review Count",
    "Buy Box: Current": "Price (¥)",
    "Buy Box: 1 day drop %": "Discount (%)",
    "One Time Coupon: Absolute": "One Time Coupon: Absolute",
    "One Time Coupon: Percentage": "One Time Coupon: Percentage",
    "One Time Coupon: Subscribe & Save %": "One Time Coupon: Subscribe & Save %",
    "Buy Box: Stock": "Buy Box: Stock",
    "Referral Fee %": "Referral Fee %",
    "Referral Fee based on current Buy Box price": "Referral Fee based on current Buy Box price (¥)",
    "ASIN": "ASIN",
}

PERCENT_COLUMNS = ["Discount (%)", "Referral Fee %"]

# === ENRICHMENT HELPERS ===
def enrich_row_local(row):
    asin = row.get("ASIN", "")
    row["Affiliate Link"] = f"https://www.amazon.co.jp/dp/{asin}/?tag={AFFILIATE_TAG}" if asin else ""
    row["Extra Discount"] = "Yes" if any([
        pd.notna(row.get("One Time Coupon: Absolute")),
        pd.notna(row.get("One Time Coupon: Percentage")),
        pd.notna(row.get("One Time Coupon: Subscribe & Save %"))
    ]) else ""
    return row

def safe_bool(val):
    if isinstance(val, list): return bool(val[-1])
    return bool(val)

def safe_float(val):
    if isinstance(val, list): return float(val[-1])
    return float(val) if val not in [None, ""] else ""

# === LOAD DATA ===
print("📄 Loading files...")
export_df = pd.read_excel(INPUT_EXPORT)
template_df = pd.read_excel(TEMPLATE_FILE)

# === MAP LOCAL COLUMNS ===
print("🔄 Mapping local columns:")
for src, tgt in tqdm(MAPPED_COLUMNS.items(), desc="Mapping columns"):
    if src in export_df.columns and tgt in template_df.columns:
        template_df[tgt] = export_df[src]
    else:
        print(f"⚠️ Missing column: '{src}' in export or '{tgt}' in template")

# === ENRICH LOCALLY ===
template_df = template_df.apply(enrich_row_local, axis=1)

# === ENRICH WITH API ===
print("🔌 Querying Keepa API...")
asins = template_df["ASIN"].dropna().astype(str).tolist()
keepa = Keepa(API_KEY)
products = keepa.query(asins, domain="JP", stats=180)

print("⚙️ Enriching ASINs from API...")
for idx, asin in enumerate(tqdm(asins, desc="Enriching ASINs")):
    product = next((p for p in products if p.get("asin") == asin), {})
    stats = product.get("stats", {})
    try:
        template_df.at[idx, "isLowest"] = safe_bool(stats.get("isLowest", False))
        template_df.at[idx, "isLowest90"] = safe_bool(stats.get("isLowest90", False))
        template_df.at[idx, "buyBox30dDropPercent"] = safe_float(stats.get("buyBox30dDropPercentage"))

        sr30 = stats.get("salesRank30")
        sr90 = stats.get("salesRank90")
        trend = "↑ Improving" if sr30 and sr90 and sr30 < sr90 else \
                "↓ Worsening" if sr30 and sr90 and sr30 > sr90 else \
                "→ Stable" if sr30 and sr90 else "N/A"
        template_df.at[idx, "Trending BSR"] = trend

        var_count = product.get("variationCount")
        if var_count is not None:
            template_df.at[idx, "Variation Simplicity"] = (
                "✅ Simple" if var_count <= 3 else
                "⚠️ Moderate" if var_count <= 10 else
                "❌ Complex"
            )
        else:
            template_df.at[idx, "Variation Simplicity"] = "N/A"

        is_l90 = safe_bool(stats.get("isLowest90", False))
        drop_pct = safe_float(stats.get("buyBox30dDropPercentage"))
        fake_flag = not is_l90 and drop_pct >= 30 if drop_pct != "" else False
        template_df.at[idx, "Fake Drop Flag"] = fake_flag

        discount = template_df.at[idx, "Discount (%)"]
        rank = template_df.at[idx, "Sales Rank"]
        grade = "F"
        if pd.notna(discount) and pd.notna(rank):
            try:
                discount = float(discount)
                rank = int(rank)
                if discount >= 50 and rank <= 5000:
                    grade = "A"
                elif discount >= 40 and rank <= 10000:
                    grade = "B"
                elif discount >= 30 and rank <= 20000:
                    grade = "C"
                elif discount >= 20 or rank <= 50000:
                    grade = "D"
            except:
                grade = "F"
        if fake_flag and grade in "ABCD":
            grade = {"A": "B", "B": "C", "C": "D", "D": "F"}[grade]
        template_df.at[idx, "Grade"] = grade

    except Exception as e:
        print(f"❌ Error enriching ASIN {asin}: {e}")

# === SAVE INITIAL FILE ===
template_df.to_excel(OUTPUT_FILE, index=False)
print(f"💾 Enrichment complete → {OUTPUT_FILE}")

# === STYLE: Apply width/height/formatting ===
print("🎨 Applying formatting...")
template_wb = openpyxl.load_workbook(TEMPLATE_FILE)
template_ws = template_wb.active
output_wb = openpyxl.load_workbook(OUTPUT_FILE)
output_ws = output_wb.active

for col_letter, col_dim in template_ws.column_dimensions.items():
    if col_letter in output_ws.column_dimensions:
        output_ws.column_dimensions[col_letter].width = col_dim.width

for row_num, row_dim in template_ws.row_dimensions.items():
    if row_num in output_ws.row_dimensions:
        output_ws.row_dimensions[row_num].height = row_dim.height

header = [cell.value for cell in output_ws[1]]
for col_name in PERCENT_COLUMNS:
    if col_name in header:
        col_idx = header.index(col_name) + 1
        for row in output_ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell.value, (float, int)):
                    cell.number_format = '0.0%'

output_wb.save(OUTPUT_FILE)
print(f"✅ Styled and saved → {OUTPUT_FILE}")
