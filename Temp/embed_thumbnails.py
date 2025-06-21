import pandas as pd
import os
from io import BytesIO
from PIL import Image
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# CONFIGURATION
INPUT_FILE = "EnrichedAffiliateSheet_Test10_NO_IMAGES.xlsx"
OUTPUT_FILE = "EnrichedAffiliateSheet_Test10_WITH_THUMBNAILS.xlsx"
THUMB_SIZE = (80, 80)

# Load data with pandas
print("Reading Excel data...")
df = pd.read_excel(INPUT_FILE)
df.to_excel(OUTPUT_FILE, index=False)

# Re-open with openpyxl
print("Embedding images...")
wb = load_workbook(OUTPUT_FILE)
ws = wb.active

# Find column indexes
columns = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
image_col = columns.get("Image")
chart_col = columns.get("Chart")

def embed_image(ws, url, row, col):
    try:
        response = requests.get(url, timeout=10)
        img = Image.open(BytesIO(response.content))
        img.thumbnail(THUMB_SIZE)

        temp_path = f"thumb_{row}_{col}.png"
        img.save(temp_path)

        xl_img = XLImage(temp_path)
        xl_img.width, xl_img.height = img.size
        ws.add_image(xl_img, f"{chr(64 + col)}{row}")

        os.remove(temp_path)
        print(f"Embedded image at row {row}, column {col}")
    except Exception as e:
        print(f"Failed at row {row}, col {col}: {e}")

for row in range(2, ws.max_row + 1):
    if image_col:
        url = ws.cell(row=row, column=image_col).value
        if isinstance(url, str) and url.startswith("http"):
            embed_image(ws, url, row, image_col)
    if chart_col:
        url = ws.cell(row=row, column=chart_col).value
        if isinstance(url, str) and url.startswith("http"):
            embed_image(ws, url, row, chart_col)

wb.save(OUTPUT_FILE)
print(f"✅ Done. Saved with thumbnails to {OUTPUT_FILE}")
