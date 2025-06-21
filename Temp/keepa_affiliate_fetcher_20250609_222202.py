
import requests
import time
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

API_KEY = "aja9g1k52mr0co8l8h3csj64io1slb3j7t6se1ic5taar1skedktm3pa95qb5862"
INPUT_FILE = "KeepaExport-2025-ProductFinder.xlsx"
OUTPUT_FILE = "EnrichedAffiliateSheet.xlsx"
MAX_TOKENS = 1260
TOKEN_RECOVERY_RATE = 20 / 60  # tokens per second

def calculate_commission_tier(ref_fee):
    if ref_fee is None:
        return "Tier 4"
    if ref_fee >= 0.12:
        return "Tier 1"
    elif ref_fee >= 0.08:
        return "Tier 2"
    elif ref_fee >= 0.05:
        return "Tier 3"
    else:
        return "Tier 4"

def load_asins_from_excel(file):
    wb = load_workbook(file)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    asin_col = header.index("ASIN")
    return [row[asin_col] for row in ws.iter_rows(min_row=2, values_only=True) if row[asin_col]]

def get_keepa_data(asins, max_tokens, recovery_rate):
    token_bucket = max_tokens
    last_request_time = time.time()
    results = []

    for asin in tqdm(asins, desc="Processing ASINs"):
        elapsed = time.time() - last_request_time
        token_bucket = min(max_tokens, token_bucket + elapsed * recovery_rate)

        if token_bucket < 5:
            sleep_time = (5 - token_bucket) / recovery_rate
            time.sleep(sleep_time)
            token_bucket += sleep_time * recovery_rate

        try:
            url = f"https://api.keepa.com/product?key={API_KEY}&domain=5&buybox=1&asin={asin}"
            response = requests.get(url)
            token_bucket -= 5
            last_request_time = time.time()

            data = response.json()
            if 'products' in data and len(data['products']) > 0:
                p = data['products'][0]
                stats = p.get("stats", {})
                ref_fee = p.get("referralFee")
                tier = calculate_commission_tier(ref_fee / 100 if ref_fee else None)
                results.append({
                    "ASIN": asin,
                    "Title": p.get("title"),
                    "Brand": p.get("brand"),
                    "Buy Box Price (¥)": stats.get("buyBoxPrice"),
                    "Referral Fee %": ref_fee / 100 if ref_fee else None,
                    "Commission Tier": tier,
                    "Image": f"https://images-na.ssl-images-amazon.com/images/I/{p.get('imagesCSV','').split(',')[0]}.jpg" if p.get('imagesCSV') else None,
                    "Chart": f"https://keepa.com/#!product/5-{asin}",
                    "Rating": p.get("rating"),
                    "Review Count": p.get("ratingCount"),
                    "Coupon": p.get("coupon"),  # May not always be available
                })
        except Exception as e:
            results.append({"ASIN": asin, "Error": str(e)})
    return results

def main():
    asins = load_asins_from_excel(INPUT_FILE)
    data = get_keepa_data(asins, MAX_TOKENS, TOKEN_RECOVERY_RATE)
    df = pd.DataFrame(data)
    df.to_excel(OUTPUT_FILE, index=False)
    print(f"Completed. Output saved to {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
