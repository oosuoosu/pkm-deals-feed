import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
from PIL import Image
import time

# === CONFIGURATION ===
API_KEY = "aja9g1k52mr0co8l8h3csj64io1slb3j7t6se1ic5taar1skedktm3pa95qb5862"
INPUT_PRODUCT_FINDER = "KeepaExport-2025-ProductFinder.xlsx"
TEMPLATE_FILE = "Final_Enriched_Affiliate_Template.xlsx"
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M")
OUTPUT_FILE = f"Final_Enriched_Affiliate_Sheet_{TIMESTAMP}.xlsx"
TRACKING_ID = "pkmsalechanne-22"

# Token limits
STARTING_TOKENS = 1260
TOKENS_PER_REQUEST = 10
RECOVERY_INTERVAL_SEC = 3  # one token every 3 seconds

# === HELPER FUNCTIONS ===
def fetch_keepa_data(asin):
    url = f"https://api.keepa.com/product?key={API_KEY}&domain=5&buybox=1&stats=180&history=0&asin={asin}"
    response = requests.get(url)
    if response.status_code == 200:
        products = response.json().get("products", [])
        return products[0] if products else None
    return None

def get_affiliate_link(asin):
    return f"https://www.amazon.co.jp/dp/{asin}/?tag={TRACKING_ID}"

def get_grade(discount, sales_rank):
    if pd.isna(discount) or pd.isna(sales_rank):
        return "F"
    if discount >= 50 and sales_rank <= 5000:
        return "A"
    elif discount >= 40 and sales_rank <= 10000:
        return "B"
    elif discount >= 30 and sales_rank <= 20000:
        return "C"
    elif discount >= 20 or sales_rank <= 50000:
        return "D"
    return "F"

def get_fake_drop_flag(product):
    is_lowest = product.get("isLowest", False)
    is_lowest90 = product.get("isLowest90", False)
    drop_percent_30d = product.get("buyBox30dDropPercent", 0)
    return not is_lowest90 and drop_percent_30d >= 30

def fetch_image(asin):
    try:
        url = f"https://images-na.ssl-images-amazon.com/images/P/{asin}.jpg"
        response = requests.get(url)
        img = Image.open(BytesIO(response.content))
        img.thumbnail((100, 100))
        return img
    except:
        return None

# === MAIN PROCESS ===
df_pf = pd.read_excel(INPUT_PRODUCT_FINDER)
wb = load_workbook(TEMPLATE_FILE)
ws = wb.active

asin_list = df_pf["ASIN"].dropna().astype(str).tolist()
starting_row = 2
tokens = STARTING_TOKENS

for index, asin in enumerate(asin_list):
    if tokens < TOKENS_PER_REQUEST:
        time.sleep(RECOVERY_INTERVAL_SEC)
        tokens += 1
    product = fetch_keepa_data(asin)
    tokens -= TOKENS_PER_REQUEST

    # Map from Product Finder
    row_data = df_pf[df_pf["ASIN"] == asin].iloc[0]
    title = row_data.get("Title")
    amazon_url = row_data.get("URL: Amazon")
    sales_rank = row_data.get("Sales Rank: Current")
    rating = row_data.get("Reviews: Rating")
    review_count = row_data.get("Reviews: Review Count")
    price = row_data.get("Buy Box: Current")
    coupon = None  # if available
    referral_fee_pct = row_data.get("Referral Fee %")
    referral_fee_yen = row_data.get("Referral Fee based on current Buy Box price")
    stock = row_data.get("Buy Box: Stock")

    discount = product.get("buyBox30dDropPercent", None) if product else None
    isLowest = product.get("isLowest", None) if product else None
    isLowest90 = product.get("isLowest90", None) if product else None
    drop_percent = product.get("buyBox30dDropPercent", None) if product else None
    seller_info = product.get("sellerId") if product else None
    commission_tier = "Tier 1" if referral_fee_pct and referral_fee_pct >= 12 else "Tier 2" if referral_fee_pct and referral_fee_pct >= 8 else "Tier 3" if referral_fee_pct and referral_fee_pct >= 5 else "Tier 4"
    trending_bsr = None
    variation_simplicity = None
    grade = get_grade(discount, sales_rank)
    fake_drop = get_fake_drop_flag(product) if product else None

    # Image & Chart
    img = fetch_image(asin)
    row = starting_row + index
    if img:
        image_path = f"/tmp/{asin}.png"
        img.save(image_path)
        img_excel = ExcelImage(image_path)
        img_excel.width = 80
        img_excel.height = 80
        ws.add_image(img_excel, f"A{row}")
    ws[f"C{row}"] = title
    ws[f"D{row}"] = get_affiliate_link(asin)
    ws[f"E{row}"] = amazon_url
    ws[f"F{row}"] = sales_rank
    ws[f"G{row}"] = rating
    ws[f"H{row}"] = review_count
    ws[f"I{row}"] = price
    ws[f"J{row}"] = discount
    ws[f"K{row}"] = isLowest
    ws[f"L{row}"] = isLowest90
    ws[f"M{row}"] = drop_percent
    ws[f"N{row}"] = coupon
    ws[f"O{row}"] = referral_fee_pct
    ws[f"P{row}"] = referral_fee_yen
    ws[f"Q{row}"] = stock
    ws[f"R{row}"] = grade
    ws[f"S{row}"] = fake_drop
    ws[f"T{row}"] = seller_info
    ws[f"U{row}"] = commission_tier
    ws[f"V{row}"] = trending_bsr
    ws[f"W{row}"] = variation_simplicity
    ws[f"X{row}"] = ""  # Marketing Approved
    ws[f"Y{row}"] = ""  # 20-sec Pitch
    ws[f"Z{row}"] = ""  # Promo Image Placeholder
    ws[f"AA{row}"] = asin

wb.save(OUTPUT_FILE)
print(f"✅ Enrichment complete. File saved as {OUTPUT_FILE}")
