
import os
import requests
import pandas as pd
from io import BytesIO
from datetime import datetime
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# === CONFIG ===
FOLDER = r"C:\Users\Philip Work\Documents\Python"
IMAGE_FOLDER = os.path.join(FOLDER, "affiliate_images")
PREFIX = "Final_Enriched_Affiliate_Sheet_ImagesCharts_"

# === Find Latest File ===
excel_files = [f for f in os.listdir(FOLDER) if f.startswith(PREFIX) and f.endswith(".xlsx")]
if not excel_files:
    raise FileNotFoundError("❌ No matching Excel files found in folder.")
latest_file = max(excel_files, key=lambda f: os.path.getmtime(os.path.join(FOLDER, f)))
INPUT_FILE = os.path.join(FOLDER, latest_file)

# === Output File Name ===
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
OUTPUT_FILE = os.path.join(FOLDER, f"Processed_Affiliate_Sheet_{timestamp}.xlsx")

# === LOAD ===
df = pd.read_excel(INPUT_FILE)
wb = load_workbook(INPUT_FILE)
ws = wb.active

# === Get Column Helpers ===
def get_col_letter_by_header(header):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            return get_column_letter(col)
    return None

def get_col_index_by_header(header):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            return col
    return None

image_col = get_col_letter_by_header("Image")
approved_col_idx = get_col_index_by_header("Marketing Approved")
asin_col_index = get_col_index_by_header("ASIN")

# === Remove Old Images in Target Columns ===
target_cols = [image_col]
remaining_images = []
removed_img_count = 0
for img in ws._images:
    if hasattr(img.anchor, '_from'):
        col_letter = get_column_letter(img.anchor._from.col + 1)
        if col_letter in target_cols:
            removed_img_count += 1
            continue
    remaining_images.append(img)
ws._images = remaining_images

# === Delete Unapproved Rows ===
deleted_row_count = 0
for row in range(ws.max_row, 1, -1):
    cell_val = ws.cell(row=row, column=approved_col_idx).value
    if str(cell_val).strip().lower() != "yes":
        ws.delete_rows(row)
        deleted_row_count += 1

# === Refresh DataFrame ===
df = pd.read_excel(INPUT_FILE)
df = df[df["Marketing Approved"].astype(str).str.lower().str.strip() == "yes"].reset_index(drop=True)

# === Prep Image Folder ===
os.makedirs(IMAGE_FOLDER, exist_ok=True)

# === Embed Images & Save to Folder ===
for idx, row in tqdm(df.iterrows(), total=len(df), desc="Processing rows"):
    asin = row.get("ASIN", "")
    if not asin or pd.isna(asin):
        continue
    row_excel = idx + 2
    ws.row_dimensions[row_excel].height = 215

    # Product Image
    img_url = row.get("Image")
    if isinstance(img_url, str) and img_url.startswith("http"):
        try:
            resp = requests.get(img_url, timeout=10)
            if resp.status_code == 200:
                img_data = resp.content
                img = XLImage(BytesIO(img_data))
                img.width = 200
                img.height = 200
                ws.add_image(img, f"{image_col}{row_excel}")

                image_path = os.path.join(IMAGE_FOLDER, f"{asin}.jpg")
                with open(image_path, "wb") as f:
                    f.write(img_data)
        except Exception as e:
            print(f"❌ Image fail for ASIN {asin}: {e}")

# Header height
ws.row_dimensions[1].height = 30

# === SAVE ===
wb.save(OUTPUT_FILE)
print(f"✅ Removed {removed_img_count} old images.")
print(f"✅ Deleted {deleted_row_count} unapproved rows.")
print(f"📁 Saved enriched file to: {OUTPUT_FILE}")
